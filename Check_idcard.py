from openpyxl import load_workbook
from icecream import ic
import os


class Validate:
    
    def __init__(self):
        # 17位系数列表
        self.multiply_lis = [7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2]
        # 校验位余数对应字典
        self.dic = {0: "1", 1: "0", 2: "X", 3: "9", 4: "8", 5: "7", 6: "6", 7: "5", 8: "4", 9: "3", 10: "2"}

    def validate(self, idNum):
        # 将字符串转列表
        id_lis = list(idNum)
        # 取前17位
        idLis = id_lis[:17]
        # 最后一位校验位
        tail_num = id_lis[17]
        # 系数运算
        result = [int(x) * int(y) for x, y in zip(self.multiply_lis, idLis)]
        # 加和取余数
        res_sum = sum(result) % 11
        # 余数字典取数
        tail = self.dic.get(res_sum)
        # 校验位比对
        if tail_num == tail:
            return True
        return False
    
    def read_file(self):
        path = os.path.join(os.path.dirname(__file__), 'id.xlsx')
        wb = load_workbook(path)
        ws = wb.active
        lis = [i.value for i in ws['A'] if i.value]
        ic(len(lis))
        res_lis = []
        for id in lis:
            if len(str(id)) == 18:
                r = self.validate(id)
            else:
                r = False
            res = '正确' if r else '错误'
            res_lis.append([id, res])
        ic(res_lis)
        ws2 = wb['结果']
        for lis in res_lis:
            print(lis)
            ws2.append(lis)
        wb.save(path)


if __name__ == "__main__":
    v = Validate()
    # v.validate('230304199812134414')
    v.read_file()
