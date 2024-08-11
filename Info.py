# -*- coding:gbk -*-
from Dict import Dict
from openpyxl import load_workbook
import requests


class SMZ:

    def __init__(self) -> None:
        self.host = 'http://10.64.2.100:30010'
        self.session = requests.Session()
        headers = {
            # 'content-type': 'application/json;charset=UTF-8',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
            "Access-Token": '9dbb7d91-8163-4d15-a24a-7a0718483757'
        }
        self.session.headers.update(headers)
        self.xl_dic = {
            '��ʿ�о���': '11',
            '˶ʿ�о���': '14',
            '��ѧ����': '21',
            '��ѧר��': '31',
            '�е�ר��': '41',
            'ְҵ����': '44',
            '����ѧУ': '47',
            '��ͨ��ѧ': '61',
            '������ѧ': '71',
            '����': '71',
            'Сѧ': '81',
            '����': '90'
        }
        self.rylb_dic = {
            '��λ��ҵ': '11',
            '����ҵ': '12',
            '�����Ը�λ����': '13',
            '���幤�̻�': '14'
        }
    
    def read_file(self, path):
        wb = load_workbook(path)
        ws = wb.active
        data = [{'name': i.value, 'id': x.value, 'phone': y.value, 'xl': z.value, 'rylb': a.value} for i, x, y, z, a in zip(ws['D'], ws['E'], ws['F'], ws['G'], ws['J']) if x.value and x.value.isalnum()]
        return data
    
    def get_user(self):
        res = self.session.get(url='http://10.64.2.100:30010/hljjy/lpleaf6-employment/api/user/info')
        print(res.status_code)
        print(res.text)

    def info(self, id):
        res = self.session.post(url=self.host + '/hljjy/lpleaf6-employment/api/aa/b/a/queryAc01Page?pageNo=1', json={"aac002": id})
        if res.status_code == 200:
            data = res.json().get('data', {})
            lis = data.get('list', [])
            if lis:
                code = lis[0].get('aac001')
                return code
        else:
            print(res.status_code)

    def get_info(self, code):
        res = self.session.post(url=self.host + '/hljjy/lpleaf6-employment/api/aa/b/a/getAc01ById', data={'aac001': code})
        if res.status_code == 200:
            data = res.json().get('data', {})
            if data:
                return data
        else:
            print(res.status_code)

    def reset_info(self, data, dic):
        if not data.get('aab302'):
            data['aab302'] = "������ʡ"
        # if not data.get('aac010'):      # ���ڵ���ַ
        #     data['aac010'] = "������ʡ�����еε�������"
        if data.get('aac010') == "������ʡ�����еε�������":
            data['aac010'] = "������ʡ�����еε���С���"
        if not data.get('aac313'):      # ��ס���ڵ�����
            data['aac313'] = "������ʡ"
        if not data.get('aae006'):      # ��ס������ַ
            data['aae006'] = "������ʡ�����еε���"
        
        id = dic.get('id')
        age = '2' if int(id[-2]) % 2 == 0 else '1'
            
        data['aac004'] = age        # �Ա�
        data['aac006'] = f"{id[6: 10]}-{id[10: 12]}-{id[12: 14]}"
        data['aac011'] = self.xl_dic.get(dic.get('xl'))       # �Ļ��̶�
        data['aae005'] = dic.get('phone')
        data['aae139'] = dic.get('phone')
        # data['aac005'] = "01"       # ����
        data['aac024'] = "13"       # ������ò
        data['aac161'] = "CHN"      # ����
        data['aac027'] = self.rylb_dic.get(dic.get('rylb'))       # ��Ա���
        data['aac016'] = "1"        # ��ҵ״̬
        data['aae007'] = '158150'   # ��������
        # print(data)
        url = self.host + '/hljjy/lpleaf6-employment/api/aa/b/a/saveAc01'
        res = self.session.post(url=url, json=data)
        print(res.status_code)
        print(res.text)
    
    def all(self, dic):
        code = smz.info(id=dic.get('id'))
        data = smz.get_info(code=code)
        smz.reset_info(data=data, dic=dic)

    def main(self):
        datas = self.read_file(path=r'F:\GSH\Other\С�������8���¾�ҵ��Աʵ����.xlsx')
        # print(datas)
        for data in datas:
            self.all(data)


if __name__ == '__main__':
    smz = SMZ()
    # smz.all({'id':'230304199812134414', 'phone':'17156865385', 'xl':'��ѧ����'})
    smz.main()

