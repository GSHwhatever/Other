import requests


class Query:

    def __init__(self) -> None:
        pass

    def query_first(self, id):
        res = requests.post(url='http://10.64.2.100:31000/server/cxjg/list', json={"searchcont": id, "type": "0"})
        if res.status_code == 200:
            data = res.json().get('data')
            d = data[0] if data else None
            if d:
                payload = {
                    "elements": [
                        {"key": "moduleId", "value": "P214-001-0111BPO"},
                        {"key": "dms", "value": "datap"},
                        {"key": "aac003", "value": d.get('aac003')},
                        {"key": "aac002", "value": d.get('aac002')},
                        {"key": "aac031In", "value": "1"},
                        {"key": "aab034", "value": d.get('aab034')},
                        {"key": "aac001", "value": d.get('aac001')},
                        {"key": "aab001", "value": ""},
                        {"key": "aae140", "value": "110"}
                    ]
                }
                return payload
    
    def query_end(self, payload):
        res = requests.post(url='http://10.64.2.100:31000/server/getListByComponent', json=payload)
        if res.status_code == 200:
            data = res.json()
            d = data.get('datap') if data else None
            if d:
                for i in d:
                    print(f'缴费年度:{i.get("aae001")}')
                    month = {'Jan': '1月', 'Feb': '2月', 'Mar': '3月', 'Apr': '4月', 'May': '5月', 'Jun': '6月', 'Jul': '7月', 'Aug': '8月', 'Sept': '9月', 'Oct': '10月', 'Nov': '11月', 'Dec': '12月'}
                    print([month.get(n) for n in month.keys() if i.get(n)!='√'])
            else:
                print(data)
        else:
            print(res.status_code)
            print(res.text)
    
    def read(self):
        from openpyxl import load_workbook
        wb = load_workbook(r'C:/Users/Administrator/Desktop/附件1_XX单位原公益性岗位人员相关信息统计表.xlsx')
        ws = wb.active
        datas = [[x.value, ''.join(y.value.split('.')[:2]), '202409'] if z.value=='至今' else [x.value, ''.join(y.value.split('.')[:2]), ''.join(z.value.split('.')[:2])] for x, y, z in zip(ws['C'], ws['D'], ws['E']) if x.value and x.value.isalnum()]
        datas.pop(0)
        for data in datas:
            print(data)
            payload = Q.query_first(data[0])
            Q.query_end(payload)


    
if __name__ == "__main__":
    Q = Query()
    payload = Q.query_first('230304197206054626')
    Q.query_end(payload)
