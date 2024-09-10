# -*- coding:gbk -*-
# 实名制导入手动校验
from openpyxl import load_workbook
import requests

class I:

    def __init__(self) -> None:
        self.host = 'http://10.64.2.100:30010'
        self.session = requests.session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
            "Access-Token": '11999f03-7567-4580-ad9a-a7df92ed2504'
        })

    def read_file(self):
        wb = load_workbook('5失业人员再就业信息明细台账.xlsx')
        ws = wb.active
        dic = {x.value: y.value for x,y in zip(ws['F'], ws['O']) if x.value and x.value.isalnum()}
        return dic

    def req(self):
        # 230304303303河北
        # 230304304302小半道
        # 230304304307立井
        # 230304304304白灰窑
        # 230304304301大半道
        dic = {
            "aae017": "230304303303",
            "acc033Query": [],
            "pageNo": 1,
            "pageSize": 110
        }
        datas = self.read_file()
        # print(datas)
        res = self.session.post(url=self.host + '/hljjy/lpleaf6-employment/api/cb/d/a/queryCc03New', json=dic)
        if res.status_code==200:
            lis = res.json().get('list', [])
            # print(lis)
            for i, d in enumerate(lis, start=1):
                gw = datas.get(d.get('aac002'))
                if gw:
                    d['aca112'] = gw
                print(f"{i}-{d.get('aac003')}-{gw}")
                res2 = self.session.post(url=self.host + '/hljjy/lpleaf6-employment/api/cb/d/a/modifyCc03New', json=d)
                print(res2.status_code)
        else:
            print(res.status_code)
            print(res.text)


if __name__ == '__main__':
    i = I()
    i.req()
    # i.read_file()
        
    