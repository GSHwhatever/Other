"""
Excel复制粘贴脚本
附带格式调整
"""

from icecream import ic
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from openpyxl.utils import get_column_letter, column_index_from_string 
from openpyxl.styles import Alignment, Border, Side, Font


class MyCopy:

    def __init__(self):
        self.wb = load_workbook('test.xlsx', data_only=True)
        self.ws = self.wb.active
    
    def my_copy(self):
        pass

    def change_style(self, ws, rows=None):
        """
        修改单元格样式(边框线、对齐方式、字体字号)
        ws: sheet表
        rows: 标题、表头行
        """
        border_style = Border(left=Side(style='thin'), 
                                    right=Side(style='thin'), 
                                    top=Side(style='thin'), 
                                    bottom=Side(style='thin'))  # 设置单元格的边框样式，全部为细线
        alignment_style = Alignment(horizontal='center', vertical='center') # 设置单元格的对齐方式，居中对齐
        font = Font(name='宋体', size=11)
        mer_lis = [mer_cell.coord.split(':')[0] for mer_cell in ws.merged_cells.ranges]
        # 修改单元格样式
        for row in ws:
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                coordinate = f'{cell.column_letter}{cell.row}'
                if coordinate in mer_lis:
                    continue
                if rows and cell.row in rows:
                    continue
                else:
                    cell.border = border_style
                    cell.alignment = alignment_style
                    cell.font = font


    def change_form(self, ws, header, headers):
        """
        修改单元格格式(行高、列宽)
        ws: sheet表
        header: 标题行
        headers: 标题、表头行
        """
        # 调整行高
        for i in [i_row for i_row, _ in enumerate(ws.rows, start=1) if i_row not in headers]:
            ws.row_dimensions[i].height = 20
        # 调整列宽
        for i_col, col in enumerate(ws.columns, start=1):
            l = [(len(str(cell.value).encode('utf-8')) + len(str(cell.value)))//2 for i_row, cell in enumerate(col, start=1) if cell.value and not isinstance(cell, MergedCell) and i_row not in (header)]
            max_width = max(l)
            ws.column_dimensions[get_column_letter(i_col)].width = max_width + 2
    
    def main(self):
        self.change_style(self.ws, [1,2])
        self.change_form(self.ws, [1], [1,2])
        self.wb.save('test2.xlsx')


if __name__ == "__main__":
    M = MyCopy()
    M.main()
    

