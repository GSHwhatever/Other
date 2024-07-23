# -*- coding:gbk -*-

from icecream import ic
import requests


class Query:

    def __init__(self) -> None:
        self.url = 'http://10.64.2.100:30010/business/m5908/entry21'        # ?aac001=&aac002=230304196508294612&aab001=&aae042=&aae041=
        self.headers = {
            'Access-Token': 'a0bfc75e-30ed-4ecc-92d2-f26e1a195f52',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
        }

    def query(self, ids, year):
        payload = {
            "aac001": "",			        # ���˱��
            "aac002": ids,	# ��ᱣ�Ϻ�
            "aab001": "",   		        # ��λ���
            "aae042": f"{year}12",		        # ��ֹ����
            "aae041": f"{year}01",		        # ��ʼ����	
            "page": "1",
            "rows": "30"
        }
        res = requests.post(url=self.url, data=payload, headers=self.headers)
        if res.status_code == 200:
            # print(res.json())
            error = res.json().get('error')
            rows = res.json().get('rows')
            if isinstance(rows, list) and not error:
                print(f'������:{res.json().get("total")}��')
                for row in rows:
                    dic = {
                        '���˱��': row.get('aac001'),
                        '����': row.get('aac003'),
                        '���֤��': row.get('aac002'),
                        '��λ���': row.get('aab001'),
                        '��λ����': row.get('aab004'),
                        '�α����': row.get('aac066'),
                        '��Ӧ�ѿ�������': row.get('aae003'),
                        '�ɷ�����': row.get('aaa115'),
                        '���ñ�־': row.get('aae792'),
                        '��������': row.get('aab191'),
                        '����ʵ�ɽ��': row.get('aae082'),
                        '������Դ': row.get('aae741')
                    }
                    ic(dic.values())
            else:
                print(f'{ids}-{error}')
        else:
            print(res.status_code)
    
    def read(self):
        from openpyxl import load_workbook
        wb = load_workbook(r'C:/Users/XBD/Desktop/������ѯ.xlsx')
        ws = wb['Sheet1']
        id_card = [i.value for i in ws['E']][2:82]
        for ids in id_card:
            self.query(ids, '2022')
    

if __name__ == '__main__':
    Q = Query()
    # Q.read()
    Q.query('230304197202034441', '2022')

