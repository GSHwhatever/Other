# -*- coding:gbk -*-
from icecream import ic
from copy import copy
from faker import Faker
from openpyxl import load_workbook
import requests, os


class TY:

    def __init__(self) -> None:
        self.url = 'https://capi.tianyancha.com/cloud-tempest/web/searchCompanyV4'
        self.path = os.path.join(os.path.dirname(__file__), '��λ.xlsx')
        self.session = requests.Session()
        # self.session.headers.update({
        #     "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.6261.95 Safari/537.36",
        #     "X-Auth-Token": "eyJhbGciOiJIUzUxMiJ9.eyJzdWIiOiIxNzE1Njg2NTM4NSIsImlhdCI6MTcyMzUzNDY4OCwiZXhwIjoxNzI2MTI2Njg4fQ.TdHIWj-5XVjZPgRSlyeRtNu452B35q0PXAzYaLwYrVBmqVpVPkhi-shvTarTWpnuBe2V_ii6fKy41VnLfiuxeg",
        #     "X-Tycid": "a41b7ad0591e11ef86be718d93a3008a"
        # })
        headers = {
            "Accept": "application/json, text/plain, */*",
            "Accept-Encoding": "gzip, deflate, br",
            "Accept-Language": "zh-CN,zh;q=0.9",
            "Connection": "keep-alive",
            "Content-Length": "1265",
            "Content-Type": "application/json",
            "Eventid": "i246",
            "Host": "capi.tianyancha.com",
            "Origin": "https://www.tianyancha.com",
            "Page_id": "SearchResult",
            "Pm": "451",
            "Referer": "https://www.tianyancha.com/nsearch?key=&sessionNo=1723534757.64312182&economicTypeMethod=1&institutionTypeMethod=1&areaCode=00230302V2020%2C00230304V2020&industryCode=61",
            "Sec-Ch-Ua": '"Chromium";v="122", "Not(A:Brand";v="24", "Google Chrome";v="122"',
            "Sec-Ch-Ua-Mobile": "?0",
            "Sec-Ch-Ua-Platform": '"Windows"',
            "Sec-Fetch-Dest": "empty",
            "Sec-Fetch-Mode": "cors",
            "Sec-Fetch-Site": "same-site",
            "Spm": "i246",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.6261.95 Safari/537.36",
            "Version": "TYC-Web",
            "X-Auth-Token": "eyJhbGciOiJIUzUxMiJ9.eyJzdWIiOiIxNzE1Njg2NTM4NSIsImlhdCI6MTcyMzYyMzU2MSwiZXhwIjoxNzI2MjE1NTYxfQ.PeBwKGtsA1zBLIDbaLiYbdwWu8OHZkWYuKJukxdQFH4KA1RbshL_cWsqbGCxKUsro9tt89TN7qryGShUl5XfeA",
            "X-Tycid": "a41b7ad0591e11ef86be718d93a3008a"
        }
        self.session.headers.update(headers)
        self.data = {
            '��λ����': '',
            'ͳһ������ô���': '',
            '��֯��������': '',
            '����������': '',
            '��ϵ��': '',
            '��ϵ�绰': '',
            'ͨѶ��ַ': '',
            '������ҵ': '���޺���ҵ����ҵ',
            '��ҵ���': '������ҵ', 
            '״̬': ''
        }
        self.faker = Faker('zh_CN')
        self.wb = load_workbook(self.path)
        self.ws = self.wb['���޺���ҵ����ҵ']
        self.bh = max([i.value if isinstance(i.value, int) else 0 for i in self.ws['A'] if i.value])

    def query_info(self, page):
        j = '{"economicTypeMethod":{"key":"economicTypeMethod","items":[{"value":"1"}]},"institutionTypeMethod":{"key":"institutionTypeMethod","items":[{"value":"1"}]},"word":{"key":"word","items":[{"value":""}]},"areaCode":{"key":"areaCode","items":[{"name":"������ʡ","value":"00230000V2020","base":"hlj","briefName":"������","childList":[{"province":"������ʡ","name":"������","value":"00230300V2020","base":"jixi","briefName":"����","childList":[{"province":"������ʡ","city":"������","name":"������","value":"00230302V2020","briefName":"����"},{"province":"������ʡ","city":"������","name":"�ε���","value":"00230304V2020","briefName":"�ε�"}]}]}]},"industryCode":{"key":"industryCode","items":[{"value":"K","name":"���ز�ҵ","childList":[{"value":"70","name":"���ز�ҵ","childList":[{"value":"709","name":"�������ز�ҵ"}]}]}]}}'
        data = {
            "allowModifyQuery": 1,
            "filterJson": j,
            "pageNum": page,
            "pageSize": 20,
            "reportInfo": {
                "distinct_id": "313345209",
                "page_id": "SearchResult",
                "page_name": "�����������ҳ",
                "search_session_id": "1723517572.33805809",
                "tab_id": "company",
                "tab_name": "��˾"
            },
            "searchType": 1,
            "sessionNo": "1723517572.33805809"
        }
        res = self.session.post(url=self.url, json=data)
        if res.status_code == 200:
            if res.json().get('state') == 'ok':
                data = res.json().get('data')
                companyList = data.get('companyList')
                if companyList:
                    ic(len(companyList))
                    self.get_info(companyList)
                else:
                    print('not companyList')
                    print(res.text)
            else:
                ic(res.json())
        else:
            print(res.status_code)

    def get_info(self, companyList):
        for dic in companyList:
            if dic.get('regStatus') == '����':
                data = copy(self.data)
                data['��λ����'] = dic.get('name')
                data['����������'] = data['��ϵ��'] = dic.get('legalPersonName')
                data['״̬'] = dic.get('regStatus')
                data['ͳһ������ô���'] = dic.get('creditCode')
                data['��֯��������'] = dic.get('orgNumber')
                data['ͨѶ��ַ'] = dic.get('regLocation')
                data['��ϵ�绰'] = dic.get('phoneInfoList', [{}])[0].get('number') if dic.get('phoneInfoList', [{}]) else self.faker.phone_number()
                ic(data)
                self.save_info(data)
            else:
                ic(dic.get('regStatus'))

    def save_info(self, data): 
        self.bh += 1
        self.ws.append([self.bh] + list(data.values()))
        self.wb.save(self.path)


if __name__ == '__main__':
    ty = TY()
    ty.query_info(page=1)
    ty.query_info(page=2)

