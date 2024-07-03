# -*- coding:gbk -*-
"""
���۲��ѯͳһ���ô������֯�ṹ����
---bak---
    //span[@class="index_copy-text__ri7W6"]/text()
    dic = {
        "���": '1',
        "��������": html.xpath('//h1[@class="index_company-name__LqKlo"]/text()'),
        "ͳһ���ô���": html.xpath('//div[@class="detail-item index_detail-item-first__IS2_h"]/div/div/span/text()'),
        "��֯��������": html.xpath('//table[@class="index_tableBox__ZadJW"]/tbody/tr[5]/td[6]/div/span/text()'),
        "��ַ": html.xpath('//div[@class="detail-item index_detail-item-second__0_Da1 index_detail-item-second-wholeline__bshTl"]/div/div/span/text()')
    }
"""
from openpyxl import load_workbook
from urllib.parse import quote
from lxml import etree
from pprint import pprint
from tqdm import tqdm
import os, requests, time, re


class TYC:
    
    def __init__(self):
        self.report_dic = {}
        self.headers = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'Accept-Encoding': 'gzip, deflate, br, zstd',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'Cookie': self.read_cookie(),
            # 'Content-Type': 'text/plain; charset=utf-8',
            'Host': 'www.tianyancha.com',
            'Pragma': 'no-cache',
            'Referer': 'https://www.tianyancha.com/search?key={}&sessionNo=1714140811.36244379',
            'Sec-Ch-Ua': '"Chromium";v="122", "Not(A:Brand";v="24", "Google Chrome";v="122"',
            'Sec-Ch-Ua-Mobile': '?0',
            'Sec-Ch-Ua-Platform': '"Windows"',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'same-origin',
            'Sec-Fetch-User': '?1',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36'
        }
        self.session = requests.Session()
        p = os.path.join(os.getcwd(), f'���۲���{time.strftime("%Y-%m-%d")}.xlsx')
        path = p if os.path.exists(p) else os.path.join(os.path.dirname(__file__), '���۲���.xlsx')
        self.wb = load_workbook(path)
    
    def read_cookie(self):
        path = os.path.join(os.path.dirname(__file__), 'cookie.txt')
        if os.path.exists(path):
            with open(path, 'r') as f:
                token = f.read()
            return token

    def req_search(self, company):
        url = f'https://www.tianyancha.com/search?key={company}'
        self.headers['Referer'] = url
        self.session.headers.update(self.headers)
        res = self.session.get(url=url)
        if res.status_code == 200:
            # print(res.text)
            html = etree.HTML(res.text)
            href = html.xpath('//a[@class="index_alink__zcia5 link-click"]/@href')
            if href:
                return href[0]
            else:
                print(f'href:{href}')
        else:
            print(f'��һ������:{res.status_code}')

    def get_code(self, url):
        res = self.session.get(url=url)
        if res.status_code == 200:
            dic = {
                "���": '1',
                "��������": '',
                "ͳһ���ô���": '',
                "��֯��������": '',
                "��ַ": ''
            }
            html = etree.HTML(res.text)
            code = html.xpath('//div[@class="index_detail-info-item__oAOqL"]/div/span/text()')
            name = html.xpath('//h1[@class="index_company-name__LqKlo"]/text()')
            print(f'code:{code}')
            if name:
                dic['��������'] = name[0]
            if code:
                dic['ͳһ���ô���'] = code[0]
                # dic['��ַ'] = code[1]
            if "��֯��������" in res.text:
                pattern = r'<td>��֯��������(.*?)</td></tr>'
                result = re.search(pattern, res.text)
                if result:
                    dic['��֯��������'] = result.group(1).split('>')[-3].split('<')[0]
            # pprint(dic, sort_dicts=False)
            return dic
        else:
            print(f'�ڶ�������:{res.status_code}')
    
    def read_file(self, path):
        if not os.path.exists(path):
            pprint(f'{path},Ŀ���ļ�������')
        else:
            wb = load_workbook(path)
            ws = wb.active
            dic = [cell.value for cell in ws['A'] if cell.value is not None]
            return dic
    
    def write_excel(self, dic):
        # ����ѯ���д��Excel�ļ�
        ws = self.wb['���۲�']
        bh = max([0 if cell.value == '���' else int(cell.value) for cell in ws['A'] if cell.value is not None])
        dic["���"] = bh + 1
        ws.append(list(dic.values()))
        
        self.wb.save(os.path.join(os.getcwd(), f'���۲���{time.strftime("%Y-%m-%d")}.xlsx'))

    def query(self, company):
        if company in self.report_dic.keys():
            dic = self.report_dic.get(company)
        else:
            com_pany = quote(company)
            r = self.req_search(com_pany)
            if r:
                dic = self.get_code(r)
                if dic:
                    self.report_dic.update({company: dic})
            else:
                print(f'r:{r}')
        self.write_excel(dic)
    
    def main(self, str):
        if os.path.isfile(str) and str.split('.')[-1] in ('xlsx', 'xlsm', 'xltx', 'xltm'):
            tag = 'file'
            d = self.read_file(str)
            print(d)
        else:
            tag = 'one'
            d = [str]
        for i in tqdm(d):
            self.query(i)
            time.sleep(1)


if __name__ == "__main__":
    tyc = TYC()
    # tyc.main('�����м�����������ӳ�')
    str = r'C:\Users\Administrator\Desktop\test.xlsx'
    tyc.main(str)
    # while True:
    #     input_str = input('������һ����������,��ֱ�Ӱ��س���ȡtest.xlsx�ļ�\n')
    #     str = os.path.join(os.getcwd(), 'test.xlsx') if not input_str else input_str
    #     tyc.main(str)
