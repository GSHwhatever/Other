"""
2025年4月17日
社保补贴根据报送数据，计算补贴起始和终止年月，校验补贴月数是否合法
"""
from openpyxl import load_workbook
from pprint import pprint


year = 2023
wb = load_workbook(r'C:\Users\WIN10\Desktop\test.xlsx', data_only=True)
target_wb = load_workbook(r'C:\Users\WIN10\Desktop\社保补贴(录入系统表式).xlsx')
ws = wb['test']
new_ws = target_wb.active


keys = ['id', 'name', 'sex', 'idcard', 'base', 'month', 'rmb', 'yl_base', 'yl_month', 'yl_rmb', 'total']
datas = []
d = len([i for i in ws['D']])
for row in range(1,d+1):
    datas.append(dict(zip(keys, [i.value for i in ws[row]])))

def check(datas):
    for data in datas:
        if data.get('sex') == '男':
            age = data.get('idcard')[6:10]
            mon = data.get('idcard')[10:12]
            if year-int(age)==55 and 12-int(mon) < data.get('month'):
                print(f"男55-{data.get('name')}-{data.get('idcard')}-{12-int(mon)}-{data.get('month')}")
            if year-int(age)==60 and int(mon) < data.get('month'):
                print(f"男60-{data.get('name')}-{data.get('idcard')}-{int(mon)}-{data.get('month')}")
        if data.get('sex') == '女':
            age = data.get('idcard')[6:10]
            mon = data.get('idcard')[10:12]
            if year-int(age)==45 and 12-int(mon) < data.get('month'):
                print(f"女45-{data.get('name')}-{data.get('idcard')}-{12-int(mon)}-{data.get('month')}")
            if year-int(age)==55 and int(mon) < data.get('month'):
                print(f"女55-{data.get('name')}-{data.get('idcard')}-{int(mon)}-{data.get('month')}")
            
        
def change(datas):
    for data in datas:
        month = data.get('month')
        year2 = int(data.get('idcard')[6:10])
        if month == 12:
            data["start_month"] = f'{year}01'
            data["end_month"] = f'{year}12'
        else:
            mon = data.get('idcard')[10:12]
            if (data.get('sex') == '男' and year-year2 == 55) or (data.get('sex') == '女' and year-year2 == 45):
                if int(mon) + 1 < 10:
                    data["start_month"] = '{}0{}'.format(year, int(mon)+1)
                else:
                    data["start_month"] = '{}{}'.format(year, int(mon)+1)
                if int(month) + int(mon) < 10:
                    data["end_month"] = '{}0{}'.format(year, int(month) + int(mon))
                else:
                    data["end_month"] = '{}{}'.format(year, int(month) + int(mon))
            elif (data.get('sex') == '男' and year-year2 == 60) or (data.get('sex') == '女' and year-year2 == 55):
                data["start_month"] = f'{year}01'
                if int(mon) + 1 < 10:
                    data["end_month"] = '{}0{}'.format(year, int(mon)+1)
                else:
                    data["end_month"] = '{}{}'.format(year, int(mon)+1)
            else:
                data["start_month"] = f'{year}01'
                if int(month) < 10:
                    data["end_month"] = '{}0{}'.format(year, month)
                else:
                    data["end_month"] = '{}{}'.format(year, month)
        print(data)
        yl = f"{float(data.get('yl_rmb')):.2f}" if data.get('yl_rmb') else None
        d = ['230304', '鸡西市滴道区', '', '1223030473127964X4', '鸡西市滴道区人力资源和社会保障服务中心', data.get('idcard'), data.get('name'), '140', f"{float(data.get('rmb')):.2f}", yl, '', '', f"{float(data.get('total')):.2f}", '', '01', data.get('start_month'), data.get('end_month')]
        new_ws.append(d)

change(datas)
target_wb.save(r'C:\Users\WIN10\Desktop\{}.xlsx'.format(year))
# check(datas)