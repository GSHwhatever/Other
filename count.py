# -*- coding:gbk -*-
from icecream import ic
from openpyxl import load_workbook
import os


class Count:

    def __init__(self) -> None:
        path = os.path.join(os.environ['USERPROFILE'], 'Desktop', '2012-2022年补贴总表.xlsx')
        self.tar_path = os.path.join(os.environ['USERPROFILE'], 'Desktop', 'target.xlsx')
        self.wb = load_workbook(path, data_only=True)
        self.tar_wb = load_workbook(self.tar_path, data_only=True)
        self.tar_ws = self.tar_wb.active
        self.id_dic = {}

    def read(self):
        # self.id_card = [i.value for i in self.tar_ws['D']]
        self.id_dic = {str(cell.value):(cell.row, cell.column) for cell in self.tar_ws['D']}
        del self.id_dic['身份证号']
        # ic(self.id_dic)

    def run(self):
        ws = self.wb.active
        for m, n, x, y, z in zip(ws['B'], ws['C'], ws['D'], ws['E'], ws['H']):
            l = [m.value, n.value, x.value, y.value, z.value]
            if all(l):
                l.append(f'2022-{z.value}')
                ic(l)
                self.tar_ws.append(l)
        self.tar_wb.save(self.tar_path)

    def run2(self):
        ws = self.wb['2020年明细']
        for m, n, x, y, z in zip(ws['B'], ws['C'], ws['D'], ws['E'], ws['H']):
            if str(y.value) in self.id_dic.keys():
                tp = self.id_dic.get(str(y.value))
                self.tar_ws.cell(row=tp[0], column=tp[1]+1, value=int(self.tar_ws.cell(row=tp[0], column=tp[1]+1).value) + int(z.value))
                self.tar_ws.cell(row=tp[0], column=tp[1]+2, value=self.tar_ws.cell(row=tp[0], column=tp[1]+2).value + f',2020-{z.value}')
            else:
                l = [m.value, n.value, x.value, str(y.value), z.value, f'2020-{z.value}']
                if all(l):
                    ic(l)
                    self.tar_ws.append(l)
        self.tar_wb.save(self.tar_path)

    def run3(self):
        tag = '2012'
        ws = self.wb[tag]
        for n, x, y, z in zip(ws['B'], ws['C'], ws['D'], ws['H']):
            if str(x.value) in self.id_dic.keys():
                tp = self.id_dic.get(str(x.value))
                self.tar_ws.cell(row=tp[0], column=tp[1]+1, value=int(self.tar_ws.cell(row=tp[0], column=tp[1]+1).value) + int(z.value))
                self.tar_ws.cell(row=tp[0], column=tp[1]+2, value=self.tar_ws.cell(row=tp[0], column=tp[1]+2).value + f',{tag}-{z.value}')
            else:
                l = ['滴道区', n.value, y.value, str(x.value), z.value, f'{tag}-{z.value}']
                if all(l):
                    ic(l)
                    self.tar_ws.append(l)
        self.tar_wb.save(self.tar_path)


if __name__ == '__main__':
    c = Count()
    c.read()
    c.run3()
