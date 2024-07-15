# -*- coding:gbk -*-
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.exceptions import InvalidFileException
from collections import Counter
from pprint import pprint
from datetime import datetime
import re

class SMZ:

    def __init__(self):
        self.header = []
        self.value_lis = []
        self.sy_values = []
        self.hyhf_list = [
            "ũ���֡�������ҵ",
            "�ɿ�ҵ",
            "����ҵ",
            "������ȼ����ˮ�������͹�Ӧҵ",
            "����ҵ",
            "��ͨ���䡢�ִ�������ҵ",
            "��Ϣ���䡢�������������ҵ",
            "����������ҵ",
            "ס�޺Ͳ���ҵ",
            "����ҵ",
            "���ز�ҵ",
            "���޺��������ҵ",
            "��ѧ�о�����������͵��ʿ���ҵ",
            "ˮ���������͹�����ʩ����ҵ",
            "����������������ҵ",
            "����",
            "��������ᱣ�Ϻ���ḣ��ҵ",
            "�Ļ�������������ҵ",
            "��������������֯",
            "������֯"
        ]
        self.cyhf_list = ['��һ��ҵ', '�ڶ���ҵ', '������ҵ']
    
    def get_headers(self, ws, min_num, max_num):
        # ��ȡexcel��ͷ
        if min_num != max_num:
            # ���ϱ�ͷ�����
            mer_row = [mer_cell.coord for mer_cell in ws.merged_cells.ranges if (min_num in (mer_cell.min_row, mer_cell.max_row) or max_num in (mer_cell.min_row, mer_cell.max_row))]
            # �õ����ϱ�ͷ�к��кϲ���Ԫ�������
            rows_list = []
            # �Զ�������(A2,AA2,ABC2)����
            for long in range(5, 20, 2):
                rows = [i for i in mer_row if len(i)==int(long)]
                if not rows:
                    break
                # ���򣬱�֤��ͷ˳�򣬷���ֱ��д��һ������
                rows.sort(key=lambda x: x.split(':')[0])
                rows_list.extend(rows)
            header = []
            for i in rows_list:
                numbers = re.findall('[0-9]+', i)
                if numbers[0] == numbers[1]:
                    # ��Ԫ��ͬ�кϲ�
                    start, end = re.findall('[A-Z]+', i)
                    # ��������ĸת��Ϊ����
                    col_num1 = column_index_from_string(start[0])
                    col_num2 = column_index_from_string(end[0])
                    for n in range(col_num1, col_num2 + 1):
                        coord = get_column_letter(n) + str(int(i[1]) + 1)
                        if coord in [i[:2] for i in mer_row]:
                            continue
                        header.append(ws[coord].value.replace('\n', '').replace(' ', '').replace('�����', ''))

                else:
                    value = ws[i.split(':')[0]].value
                    if value:
                        header.append(value.replace('\n', '').replace(' ', '').replace('�����', ''))
            h2 = [i.value for i in ws[max_num]]
            if len(header) != len(h2):
                l = h2[len(header) - len(h2):]
                if None not in l:
                    header.extend(l)
        else:
            # ���б�ͷֱ��ȡ
            header = [i.value.replace('\n', '').replace(' ', '').replace('�����', '') for i in ws[max_num] if i.value]
        # print(header)
        return header
    
    def syry(self, ws_sy):
        header_sy = self.get_headers(ws_sy, 2, 2)
        syry_values = []
        for row in [i for i in ws_sy.iter_rows()][2:]:
            if not row[0].value:
                break
            lis_sy = []
            for i in row:
                lis_sy.append(i.value)
            syry_values.append(dict(zip(header_sy, lis_sy)))
        keys = ['���', '����', '�Ա�', '����', '�Ļ��̶�', '���֤��', '��������', '�����س�', '��ҵ��ҵ֤��', '�Ƿ�Ǽ�ʧҵ��Ա', 'ʧҵ��Ա����', 'ʧҵʱ��', '��ȡʧҵ���ս���ֹʱ��', '��ְ����', '��ѵ����', '��ҵ��������', '��ϵ�绰', '����', '�ȼ�']
        for i in syry_values:
            values = []
            for key in keys:
                v = i.get(key)
                if key == '���':
                    v = i.get('�����')
                if key == '����':
                    v = datetime.now().year - int(i.get('���֤��')[6:10])
                if key == '�Ļ��̶�':
                    v = i.get('ѧ��')
                if key == '��������':
                    v = '����'
                if key == '�����س�':
                    v = i.get('���⼼��')
                if key == '�Ƿ�Ǽ�ʧҵ��Ա':
                    v = '��'
                if key == 'ʧҵ��Ա����':
                    v = '(9)'
                if key == '��ȡʧҵ���ս���ֹʱ��':
                    v = '��'
                if key == '��ְ����':
                    v = '����'
                if key == '��ѵ����':
                    v = '��'
                if key == '��ҵ��������':
                    v = '��1��'
                if key == '��ϵ�绰':
                    v = i.get('�绰')
                if key == '����':
                    v = '�ж���'
                if key == '�ȼ�':
                    v = 'ȱ������' if i.get('���⼼��') == '��' else '�߱�����'
                if key == 'ʧҵʱ��':
                    v_time = i.get('ʧҵʱ��')
                    if v_time:
                        month = datetime.now().month - v.month
                        v = v_time.replace(month=datetime.now().month - 2) if month > 2 else v_time
                    else:
                        v = datetime.now().replace(month=datetime.now().month - 2, day=1, hour=0, minute=0, second=0, microsecond=0)
                values.append(v)
            self.sy_values.append(values)

    def read_file(self, file, min_num, max_num):
        # ��excel�ļ���ȡ���ݣ�����ͷ����������ֵ䱣����self.value_lis�б���
        # path = os.path.join(os.path.dirname(__file__), '��־�ҵ��������̨��', file)
        path = file
        # print(f'path:{path}')
        self.value_lis = []
        try:
            wb = load_workbook(path, data_only=True)
        except InvalidFileException as E:
            print(f'���ܴ���{path.split(".")[-1]}���͵�Excel�ļ�,�Ƽ����Ϊxlsx����')
            return 'error'
        else:
            if "ʵ����" in path.split('/')[-1] and 'ʧҵ��Ա���' in wb.sheetnames:
                self.syry(wb['ʧҵ��Ա���'])
            try:
                ws = wb['�¾�ҵ��Ա']
            except KeyError:
                ws = wb.worksheets[0]
            finally:
                header = self.get_headers(ws, min_num, max_num)
                for row in [i for i in ws.iter_rows()][max_num:]:
                    if not row[0].value:
                        break
                    lis = []
                    for i in row:
                        lis.append(i.value)
                    self.value_lis.append(dict(zip(header, lis)))

    def get_message(self, datas):
        # ��ȡ̨�˾�����Ϣ
        syzjy = [i for i in datas if i.get('��ҵ����') == 'ʧҵ�پ�ҵ']
        jyknzjy = [i for i in datas if i.get('��ҵ������Ա') == '��']
        dzysxl = [i for i in datas if i.get('ѧ��') in ('��ѧר��', '��ѧ����', '˶ʿ�о���')]
        woman = [i for i in datas if i.get('�Ա�') == 'Ů']
        rs1624 = [i for i in datas if int(i.get('����')) in range(16, 25)]
        rs2545 = [i for i in datas if int(i.get('����')) in range(25, 46)]
        rs4660 = [i for i in datas if int(i.get('����')) in range(46, 61)]
        counts = Counter([i.get('���²�ҵ����') for i in datas])
        cyhf = [counts.get(industry, 0) for industry in self.cyhf_list]
        counts = Counter([i.get('������ҵ(��ҵ��ʽΪ��λ��ҵʱ����)') for i in datas])
        hyhf = [counts.get(industry, 0) for industry in self.hyhf_list]
        dwjy = [i for i in datas if i.get('��ҵ��ʽ') == '��λ��ҵ']
        lhjy = [i for i in datas if i.get('��ҵ��ʽ') == '����ҵ']
        # print(hyhf)
        dic = {
            "�¾�ҵ����": len(datas),
            "ʧҵ��Ա�پ�ҵ����": len(syzjy),
            "��ҵ�����پ�ҵ����": len(jyknzjy),
            "��ר����ѧ������": len(dzysxl),
            "Ů������": len(woman),
            "�ǹ�������ҵ": len(dwjy),
            "����ҵ": len(lhjy),
            "16-24��������": len(rs1624),
            "25-45��������": len(rs2545),
            "46-60��������": len(rs4660),
            "��ҵ����": dict(zip(self.cyhf_list, cyhf)),
            "��ҵ����": dict(zip(self.hyhf_list, hyhf))
        }
        return dic
    
    def get_all_message(self, path):
        self.read_file(path, 2, 3)
        # ic(self.value_lis)
        dic = self.get_message(self.value_lis)
        pprint('�¾�ҵ��Ա��Ϣ��')
        pprint(dic)
        pprint(f'ʧҵ��Ա������{len(self.sy_values)}')
        pprint(f'����Ů��������{len([i for i in self.sy_values if "Ů" in i])}')


if __name__ == "__main__":
    smz = SMZ()
    smz.get_all_message(r'D:/Win7����/D/��΢��/���2024/7��/С�������7���¾�ҵ��Աʵ����.xlsx')
